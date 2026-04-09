"""Report generation service - DOCX/PPTX via Markpresso API, XLSX via openpyxl."""

import io
import httpx
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from data.constants import (
    SITES, TGBTS, CONSUMPTION_BREAKDOWN, PV_INSTALLATION,
    MONTHLY_PV_VS_CONSUMPTION, RECOMMENDATIONS, PV_ECONOMICS,
    DAY_NIGHT, CO2_PROJECTIONS,
)
from services.ai_service import (
    generate_executive_summary, generate_consumption_analysis,
    generate_recommendations, generate_pv_assessment,
)

MARKPRESSO_URL = "https://markpresso.pro/api/v1/convert"

MARKPRESSO_STYLE = {
    "primaryColor": "#7CB342",
    "fontHeading": "Arial",
    "fontBody": "Arial",
    "companyName": "Sungy SPA",
}


def _build_report_markdown(req) -> str:
    """Build the full audit report as Markdown."""
    site = SITES.get(req.site_id, SITES["bel-kolea"])
    tgbts = TGBTS.get(req.site_id, TGBTS["bel-kolea"])
    pv = PV_INSTALLATION
    cb = CONSUMPTION_BREAKDOWN
    dn = DAY_NIGHT
    now = datetime.now().strftime("%B %Y")

    sections = []

    # --- Title ---
    sections.append(f"# Rapport d'Audit Energetique\n")
    sections.append(f"**{site['name']}**  ")
    sections.append(f"{site['address']}  ")
    sections.append(f"Date du rapport : {now}  ")
    sections.append(f"Redige par **Sungy SPA**\n")
    sections.append("---\n")

    # --- AI Executive Summary ---
    sections.append("## 1. Synthese Executive\n")
    if req.include_ai:
        summary = generate_executive_summary(site)
        sections.append(f"{summary}\n")
    else:
        sections.append("*(Synthese IA desactivee)*\n")

    # --- Site Presentation ---
    sections.append("## 2. Presentation du Site\n")
    sections.append("| Information | Detail |")
    sections.append("|---|---|")
    sections.append(f"| Entreprise | {site['name']} |")
    sections.append(f"| Activite | {site['activity']} |")
    sections.append(f"| Adresse | {site['address']} |")
    sections.append(f"| Tarification | {site['contract_type']} |")
    sections.append(f"| Puissance souscrite | {site['subscribed_power_kw']} kW |")
    sections.append(f"| Consommation annuelle | {site['annual_consumption_mwh']} MWh |")
    sections.append(f"| Zone climatique | {site.get('climate_zone', 'A')} |")
    sections.append(f"| Temperature moyenne | {site.get('avg_temp_c', 18.2)} C |\n")

    # --- Electrical Analysis ---
    sections.append("## 3. Analyse Electrique\n")

    if req.include_ai:
        analysis = generate_consumption_analysis(tgbts)
        sections.append(f"{analysis}\n")

    # TGBT details table
    sections.append("### 3.1 Detail par TGBT\n")
    sections.append("| TGBT | S moy (kVA) | S max (kVA) | S transfo (kVA) | Taux charge | Cos phi | Part |")
    sections.append("|---|---|---|---|---|---|---|")
    for t in tgbts:
        sections.append(
            f"| {t['name']} | {t['s_moy_kva']} | {t['s_max_kva']} | {t['s_transfo_kva']} "
            f"| {t['tc_pct']}% | {t['cos_phi']} | {t['share_pct']}% |"
        )
    sections.append("")

    # Tariff analysis
    sections.append("### 3.2 Analyse Tarifaire\n")
    sections.append("| Periode | Horaires | Tarif (DZD/kWh) | Consommation (MWh) | Cout (DZD) |")
    sections.append("|---|---|---|---|---|")
    sections.append(f"| **Pointe** | 17:00-21:00 | 8,72 | {cb['pointe']['mwh']} | {cb['pointe']['cost_dzd']:,} |")
    sections.append(f"| **Pleine** | 06:00-17:00 + 21:00-22:30 | 1,94 | {cb['pleine']['mwh']} | {cb['pleine']['cost_dzd']:,} |")
    sections.append(f"| **Creuse** | 22:30-06:00 | 1,02 | {cb['creuse']['mwh']} | {cb['creuse']['cost_dzd']:,} |")
    sections.append(f"| **Total** | - | - | {cb['total']['mwh']} | {cb['total']['cost_dzd']:,} |")
    sections.append("")
    sections.append("> **Constat important :** Les heures de pointe representent seulement 14% de la consommation mais 46% de la facturation.\n")

    # Day/Night
    sections.append("### 3.3 Ventilation Jour/Nuit\n")
    sections.append(f"- **Jour** ({dn['day_hours']}) : **{dn['day_pct']}%** de la consommation")
    sections.append(f"- **Nuit** ({dn['night_hours']}) : **{dn['night_pct']}%** de la consommation")
    sections.append(f"\nLa consommation nocturne elevee est principalement due au fonctionnement continu des groupes froids.\n")

    # --- Recommendations ---
    sections.append("## 4. Recommandations de Sobriete Energetique\n")

    if req.include_ai:
        recs = generate_recommendations(site)
        sections.append(f"{recs}\n")

    sections.append("### Recapitulatif des recommandations\n")
    sections.append("| Recommandation | Economie energie (MWh/an) | Investissement (DZD) | CO2 evite (t/an) | Economies (DZD/an) |")
    sections.append("|---|---|---|---|---|")
    for r in RECOMMENDATIONS:
        sections.append(
            f"| {r['description']} | {r['energy_saving_mwh']} | {r['investment_dzd']:,} "
            f"| {r['co2_avoided_tonnes']} | **{r['savings_dzd']:,}** |"
        )
    sections.append("")

    # --- PV Dimensioning ---
    sections.append("## 5. Dimensionnement Photovoltaique\n")

    if req.include_ai:
        pv_text = generate_pv_assessment(pv)
        sections.append(f"{pv_text}\n")

    sections.append("### Specifications de l'installation\n")
    sections.append(f"- **Puissance** : {pv['capacity_kwc']} kWc")
    sections.append(f"- **Panneaux** : {pv['panel_count']} x {pv['panel_type']}")
    sections.append(f"- **Production annuelle** : {pv['annual_production_mwh']} MWh")
    sections.append(f"- **Taux de couverture** : {pv['coverage_rate_pct']}%")
    sections.append(f"- **Rendement specifique** : {pv['specific_yield_kwh_kwc']} kWh/kWc\n")

    sections.append("### Production mensuelle\n")
    sections.append("| Mois | Production PV (MWh) | Consommation (MWh) | Couverture |")
    sections.append("|---|---|---|---|")
    for m in MONTHLY_PV_VS_CONSUMPTION:
        cov = round(m["pv_mwh"] / m["consumption_mwh"] * 100, 1)
        sections.append(f"| {m['month']} | {m['pv_mwh']} | {m['consumption_mwh']} | {cov}% |")
    sections.append(f"| **Total** | **{pv['annual_production_mwh']}** | **6 666** | **{pv['coverage_rate_pct']}%** |")
    sections.append("")

    # Economics
    sections.append("### Retour sur investissement\n")
    sections.append(f"- **Investissement initial** : {pv['investment_dzd']:,} DZD")
    sections.append(f"- **Maintenance annuelle** : {pv['maintenance_dzd_year']:,} DZD")
    sections.append(f"- **Economies annuelles** : {PV_ECONOMICS['savings_dzd']:,} DZD")
    sections.append(f"- **ROI** : {pv['roi_years']} ans")
    sections.append(f"- **TRI** : {pv['tri_pct']}%\n")

    # CO2
    sections.append("### Impact environnemental\n")
    sections.append("| Annee | CO2 evite (tonnes) |")
    sections.append("|---|---|")
    for p in CO2_PROJECTIONS:
        sections.append(f"| Annee {p['year']} | {p['co2_tonnes']:,} |")
    sections.append("")

    return "\n".join(sections)


def _build_presentation_markdown(req) -> str:
    """Build presentation content as Markdown (each ## = new slide)."""
    site = SITES.get(req.site_id, SITES["bel-kolea"])
    tgbts = TGBTS.get(req.site_id, TGBTS["bel-kolea"])
    pv = PV_INSTALLATION
    cb = CONSUMPTION_BREAKDOWN
    now = datetime.now().strftime("%B %Y")

    slides = []

    # Slide 1: Title
    slides.append(f"# Rapport d'Analyse Electrique\n")
    slides.append(f"**{site['name']}**  ")
    slides.append(f"{site['address']}  ")
    slides.append(f"{now}\n")

    # Slide 2: Synthese
    slides.append("## Synthese Generale\n")
    slides.append(f"- **Consommation annuelle** : {site['annual_consumption_mwh']} MWh")
    slides.append(f"- **Tarification** : {site['contract_type']}")
    slides.append(f"- **Puissance souscrite** : {site['subscribed_power_kw']} kW")
    slides.append(f"- **Production PV** : {pv['annual_production_mwh']} MWh/an")
    slides.append(f"- **Taux de couverture** : {pv['coverage_rate_pct']}%")
    slides.append(f"- **CO2 evite** : {pv['co2_avoided_tonnes_year']} tonnes/an\n")

    # Slide 3: TGBT
    slides.append("## Repartition des Charges\n")
    slides.append("| TGBT | S moy (kVA) | Taux charge | Cos phi | Part |")
    slides.append("|---|---|---|---|---|")
    for t in tgbts:
        slides.append(f"| {t['name']} | {t['s_moy_kva']} | {t['tc_pct']}% | {t['cos_phi']} | {t['share_pct']}% |")
    slides.append("")

    # Slide 4: Consommation
    slides.append("## Analyse de la Consommation\n")
    slides.append(f"- Consommation enregistree : **344 MWh** (18 jours)")
    slides.append(f"- Jour : **{cb.get('total', {}).get('mwh', 6666)}** MWh/an")
    slides.append(f"- Ventilation : Jour **43%** / Nuit **57%**")
    slides.append(f"- Weekend baseline : ~500 kW (groupes froids)")
    slides.append(f"- Max journalier : 24,02 MWh — Min : 10,17 MWh\n")

    # Slide 5: Tarification
    slides.append("## Analyse Tarifaire\n")
    slides.append("| Periode | Consommation | Cout |")
    slides.append("|---|---|---|")
    slides.append(f"| Creuse (22:30-6:00) | {cb['creuse']['mwh']} MWh | {cb['creuse']['cost_dzd']:,} DZD |")
    slides.append(f"| Pleine (6:00-17:00) | {cb['pleine']['mwh']} MWh | {cb['pleine']['cost_dzd']:,} DZD |")
    slides.append(f"| Pointe (17:00-21:00) | {cb['pointe']['mwh']} MWh | {cb['pointe']['cost_dzd']:,} DZD |")
    slides.append(f"\n> Pointe = 14% conso mais 46% facture\n")

    # Slide 6: Recommandations
    slides.append("## Recommandations\n")
    for r in RECOMMENDATIONS[:5]:
        slides.append(f"- {r['description']} → **{r['savings_dzd']:,} DZD/an**")
    slides.append("")

    # Slide 7: PV
    slides.append("## Dimensionnement Photovoltaique\n")
    slides.append(f"- **{pv['capacity_kwc']} kWc** — {pv['panel_count']} panneaux")
    slides.append(f"- Production : **{pv['annual_production_mwh']} MWh/an**")
    slides.append(f"- Couverture : **{pv['coverage_rate_pct']}%**")
    slides.append(f"- CO2 evite : **{pv['co2_avoided_tonnes_year']} tonnes/an**")
    slides.append(f"- Investissement : **{pv['investment_dzd']:,} DZD**")
    slides.append(f"- ROI : **{pv['roi_years']} ans**\n")

    # Slide 8: Production mensuelle
    slides.append("## Production PV vs Consommation\n")
    slides.append("| Mois | PV (MWh) | Conso (MWh) |")
    slides.append("|---|---|---|")
    for m in MONTHLY_PV_VS_CONSUMPTION:
        slides.append(f"| {m['month']} | {m['pv_mwh']} | {m['consumption_mwh']} |")
    slides.append("")

    # Slide 9: AI Analysis
    if req.include_ai:
        slides.append("## Analyse IA\n")
        ai_text = generate_executive_summary(site)
        slides.append(f"{ai_text}\n")

    return "\n".join(slides)


async def _call_markpresso(markdown: str, title: str, fmt: str) -> bytes:
    """Call Markpresso API to convert markdown to DOCX or PPTX."""
    endpoint = f"{MARKPRESSO_URL}/{fmt}"
    payload = {
        "markdown": markdown,
        "title": title,
        "style": MARKPRESSO_STYLE,
    }
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(endpoint, json=payload)
        response.raise_for_status()
        return response.content


async def generate_docx(req) -> bytes:
    site = SITES.get(req.site_id, SITES["bel-kolea"])
    markdown = _build_report_markdown(req)
    title = f"Audit Energetique - {site['name']}"
    return await _call_markpresso(markdown, title, "docx")


async def generate_pptx(req) -> bytes:
    site = SITES.get(req.site_id, SITES["bel-kolea"])
    markdown = _build_presentation_markdown(req)
    title = f"Rapport Analyse - {site['name']}"
    return await _call_markpresso(markdown, title, "pptx")


async def generate_xlsx(req) -> bytes:
    """XLSX stays local via openpyxl (Markpresso doesn't support xlsx)."""
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="F59E0B")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    site = SITES.get(req.site_id, SITES["bel-kolea"])

    # Sheet 1: Synthese
    ws = wb.active
    ws.title = "Synthese"
    kpis = [
        ("Site", site["name"]),
        ("Adresse", site["address"]),
        ("Consommation annuelle (MWh)", site["annual_consumption_mwh"]),
        ("Tarification", site["contract_type"]),
        ("Puissance souscrite (kW)", site["subscribed_power_kw"]),
        ("Production PV (MWh/an)", PV_INSTALLATION["annual_production_mwh"]),
        ("Taux couverture PV (%)", PV_INSTALLATION["coverage_rate_pct"]),
        ("CO2 evite (t/an)", PV_INSTALLATION["co2_avoided_tonnes_year"]),
        ("Economies annuelles (DZD)", PV_ECONOMICS["savings_dzd"]),
        ("Investissement PV (DZD)", PV_INSTALLATION["investment_dzd"]),
        ("ROI (ans)", PV_INSTALLATION["roi_years"]),
    ]
    for i, (label, value) in enumerate(kpis, 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    # Sheet 2: TGBT
    ws2 = wb.create_sheet("TGBT")
    headers = ["TGBT", "S moy (kVA)", "S max (kVA)", "S min (kVA)", "S transfo (kVA)", "TC %", "Cos phi", "Part %"]
    for j, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for i, t in enumerate(TGBTS.get(req.site_id, TGBTS["bel-kolea"]), 2):
        ws2.cell(row=i, column=1, value=t["name"])
        ws2.cell(row=i, column=2, value=t["s_moy_kva"])
        ws2.cell(row=i, column=3, value=t["s_max_kva"])
        ws2.cell(row=i, column=4, value=t["s_min_kva"])
        ws2.cell(row=i, column=5, value=t["s_transfo_kva"])
        ws2.cell(row=i, column=6, value=t["tc_pct"])
        ws2.cell(row=i, column=7, value=t["cos_phi"])
        ws2.cell(row=i, column=8, value=t["share_pct"])

    # Sheet 3: Production PV
    ws3 = wb.create_sheet("Production PV")
    for j, h in enumerate(["Mois", "Production PV (MWh)", "Consommation (MWh)", "Taux couverture (%)"], 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for i, m in enumerate(MONTHLY_PV_VS_CONSUMPTION, 2):
        ws3.cell(row=i, column=1, value=m["month"])
        ws3.cell(row=i, column=2, value=m["pv_mwh"])
        ws3.cell(row=i, column=3, value=m["consumption_mwh"])
        ws3.cell(row=i, column=4, value=round(m["pv_mwh"] / m["consumption_mwh"] * 100, 1))

    # Sheet 4: Recommandations
    ws4 = wb.create_sheet("Recommandations")
    for j, h in enumerate(["Recommandation", "Economie energie (MWh/an)", "Investissement (DZD)", "CO2 evite (t/an)", "Economies (DZD/an)"], 1):
        cell = ws4.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for i, r in enumerate(RECOMMENDATIONS, 2):
        ws4.cell(row=i, column=1, value=r["description"])
        ws4.cell(row=i, column=2, value=r["energy_saving_mwh"])
        ws4.cell(row=i, column=3, value=r["investment_dzd"])
        ws4.cell(row=i, column=4, value=r["co2_avoided_tonnes"])
        ws4.cell(row=i, column=5, value=r["savings_dzd"])
    ws4.column_dimensions["A"].width = 60

    # Sheet 5: Tarification
    ws5 = wb.create_sheet("Tarification")
    for j, h in enumerate(["Periode", "Heures", "Tarif (DZD/kWh)", "Consommation (MWh)", "Cout (DZD)"], 1):
        cell = ws5.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
    cb = CONSUMPTION_BREAKDOWN
    rows = [
        ("Creuse", "22:30-06:00", 1.02, cb["creuse"]["mwh"], cb["creuse"]["cost_dzd"]),
        ("Pleine", "06:00-17:00 + 21:00-22:30", 1.94, cb["pleine"]["mwh"], cb["pleine"]["cost_dzd"]),
        ("Pointe", "17:00-21:00", 8.72, cb["pointe"]["mwh"], cb["pointe"]["cost_dzd"]),
        ("Total", "-", "-", cb["total"]["mwh"], cb["total"]["cost_dzd"]),
    ]
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws5.cell(row=i, column=j, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
